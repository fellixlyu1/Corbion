import openpyxl
import win32com
from datetime import datetime, timedelta

def check_expiration_dates(file_path_param):

    # Load the workbook and the sheet
    workbook = openpyxl.load_workbook(file_path_param)
    sheet = workbook['Sheet1']  # Assuming the expiration dates are in 'Sheet1'

    # Get today's date
    today = datetime.today()

    # Loop through the rows and check the expiration date (assuming the date is in the 4th column, i.e., column D)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):  # Assuming the first row is headers
        material = row[0].value  # Assuming material name is in the first column
        quantity = row[1].value  # Assuming quantity is in the second column
        size = row[2].value  # Assuming size is in the third column
        expiration_date = row[3].value  # Expiration date is in the 4th column

        # Check if expiration_date is a valid date
        if isinstance(expiration_date, datetime):
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            accounts = namespace.Accounts
            mail = outlook.CreateItem(0)

            # Calculate the difference in days between the expiration date and today's date
            days_left = (expiration_date - today).days

            # If the expiration date is within 30 days
            if days_left <= 30:
                subject = f"Warning: {material} Expiring Soon"
                body = (f"Hello,\n\n"
                        f"The following material is expiring in {days_left} days:\n"
                        f"Material: {material}\n"
                        f"Size: {size}\n"
                        f"Quantity: {quantity}\n"
                        f"Expiration Date: {expiration_date.date()}\n\n"
                        "Please take necessary actions.\n\n"
                        "Best regards,\n"
                        "Your Expiration Tracker System")

                # Set up the email details
                mail.Subject = subject
                mail.Body = body

                # Set recipient email (Replace with the desired recipient email)
                mail.To = "FinishedGoods@corbion.com"  # Replace with actual recipient email

                try:
                    # Send the email
                    mail.Send()
                    print(f"Email sent for {material}, expiring in {days_left} days.")
                except Exception as e:
                    print(f"Failed to send email for {material}. Error: {e}")
            else:
                print(f"Invalid expiration date for {material}.")

file_path = "Expiration_Tracker.xlsx"
check_expiration_dates(file_path)